#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI模型调用脚本
支持断点续传、进度显示和配置化管理
支持文本和图片输入（兼容视觉模型）
支持多Provider和多种API调用方式
支持Excel文件输入，自动提取嵌入图片
"""

import pandas as pd
import requests
import json
import yaml
import time
import os
import sys
import base64
import mimetypes
import io
import zipfile
import re
from typing import Dict, Any, Optional, Tuple, List, Union
from tqdm import tqdm
import argparse
import logging
from datetime import datetime
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

# 尝试导入openpyxl用于处理Excel文件
try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelImageExtractor:
    """从Excel文件中提取嵌入的图片"""
    
    def __init__(self, excel_path: str):
        """
        初始化图片提取器
        
        Args:
            excel_path: Excel文件路径
        """
        self.excel_path = excel_path
        self.images: Dict[str, str] = {}  # 单元格位置 -> Base64图片数据
        self._extract_images()
    
    def _extract_images(self):
        """从Excel文件中提取所有图片"""
        if not HAS_OPENPYXL:
            return
        
        try:
            # 方法1: 使用openpyxl提取图片
            self._extract_with_openpyxl()
            
            # 方法2: 如果openpyxl没提取到，尝试直接从xlsx解压提取
            if not self.images:
                self._extract_from_xlsx_archive()
                
        except Exception as e:
            print(f"⚠️ 提取图片时出错: {str(e)}")
    
    def _extract_with_openpyxl(self):
        """使用openpyxl提取图片"""
        try:
            wb = load_workbook(self.excel_path)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # 遍历sheet中的所有图片
                for image in sheet._images:
                    try:
                        # 获取图片锚点位置
                        anchor = image.anchor
                        if hasattr(anchor, '_from'):
                            row = anchor._from.row + 1  # openpyxl是0-indexed
                            col = anchor._from.col + 1
                        elif hasattr(anchor, 'anchortype'):
                            # TwoCellAnchor 或 OneCellAnchor
                            if hasattr(anchor, '_from'):
                                row = anchor._from.row + 1
                                col = anchor._from.col + 1
                            else:
                                continue
                        else:
                            continue
                        
                        # 读取图片数据
                        if hasattr(image, '_data'):
                            image_data = image._data()
                        elif hasattr(image, 'ref'):
                            # 从文件引用读取
                            image_data = image.ref.getvalue() if hasattr(image.ref, 'getvalue') else None
                        else:
                            continue
                        
                        if image_data:
                            # 确定MIME类型
                            mime_type = self._detect_image_mime(image_data)
                            base64_data = base64.b64encode(image_data).decode('utf-8')
                            
                            # 存储: 使用行号作为key (便于后续匹配)
                            cell_key = f"{row}"
                            self.images[cell_key] = f"data:{mime_type};base64,{base64_data}"
                            
                    except Exception as e:
                        continue
                        
            wb.close()
            
        except Exception as e:
            pass
    
    def _extract_from_xlsx_archive(self):
        """直接从xlsx文件作为zip解压提取图片"""
        try:
            # xlsx本质是一个zip文件
            with zipfile.ZipFile(self.excel_path, 'r') as zf:
                # 查找所有图片文件
                image_files = [f for f in zf.namelist() if f.startswith('xl/media/')]
                
                # 读取drawing关系文件，找到图片与单元格的对应关系
                drawing_rels = {}
                for name in zf.namelist():
                    if 'drawings/_rels' in name and name.endswith('.rels'):
                        try:
                            rels_content = zf.read(name).decode('utf-8')
                            # 解析关系文件
                            for match in re.finditer(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_content):
                                rel_id = match.group(1)
                                target = match.group(2).replace('../media/', '')
                                drawing_rels[rel_id] = target
                        except:
                            pass
                
                # 读取drawing文件，找到图片位置
                image_positions = {}
                for name in zf.namelist():
                    if 'drawings/drawing' in name and name.endswith('.xml'):
                        try:
                            drawing_content = zf.read(name).decode('utf-8')
                            
                            # 查找所有anchor及其关联的图片
                            # 匹配 <xdr:from><xdr:col>X</xdr:col><xdr:row>Y</xdr:row>
                            anchor_pattern = re.compile(
                                r'<xdr:from>\s*<xdr:col>(\d+)</xdr:col>[^<]*<xdr:colOff>[^<]*</xdr:colOff>\s*'
                                r'<xdr:row>(\d+)</xdr:row>.*?<a:blip[^>]*r:embed="(rId\d+)"',
                                re.DOTALL
                            )
                            
                            for match in anchor_pattern.finditer(drawing_content):
                                col = int(match.group(1)) + 1  # 转为1-indexed
                                row = int(match.group(2)) + 1
                                rel_id = match.group(3)
                                
                                if rel_id in drawing_rels:
                                    image_name = drawing_rels[rel_id]
                                    image_positions[f"xl/media/{image_name}"] = row
                                    
                        except Exception as e:
                            pass
                
                # 读取并编码图片
                for image_file in image_files:
                    try:
                        image_data = zf.read(image_file)
                        mime_type = self._detect_image_mime(image_data)
                        base64_data = base64.b64encode(image_data).decode('utf-8')
                        
                        # 如果有位置信息，使用行号作为key
                        if image_file in image_positions:
                            row = image_positions[image_file]
                            cell_key = f"{row}"
                        else:
                            # 否则使用文件名中的数字
                            match = re.search(r'image(\d+)', image_file)
                            if match:
                                # 假设图片按顺序对应行号（从第2行开始，第1行是标题）
                                image_num = int(match.group(1))
                                cell_key = f"{image_num + 1}"  # +1 因为标题行
                            else:
                                continue
                        
                        self.images[cell_key] = f"data:{mime_type};base64,{base64_data}"
                        
                    except Exception as e:
                        continue
                        
        except Exception as e:
            pass
    
    def _detect_image_mime(self, image_data: bytes) -> str:
        """检测图片的MIME类型"""
        # 检查文件头
        if image_data[:8] == b'\x89PNG\r\n\x1a\n':
            return 'image/png'
        elif image_data[:2] == b'\xff\xd8':
            return 'image/jpeg'
        elif image_data[:6] in (b'GIF87a', b'GIF89a'):
            return 'image/gif'
        elif image_data[:4] == b'RIFF' and image_data[8:12] == b'WEBP':
            return 'image/webp'
        else:
            return 'image/png'  # 默认PNG
    
    def get_image_base64(self, row: int) -> Optional[str]:
        """
        获取指定行的图片Base64数据
        
        Args:
            row: 行号（1-indexed，与Excel行号一致）
            
        Returns:
            Base64编码的图片数据URL，如果没有图片则返回None
        """
        return self.images.get(str(row))
    
    def get_image_base64_raw(self, row: int) -> Optional[Tuple[str, str]]:
        """
        获取指定行图片的原始Base64数据和MIME类型
        
        Args:
            row: 行号（1-indexed）
            
        Returns:
            (base64_data, mime_type) 元组，如果没有图片则返回None
        """
        data_url = self.images.get(str(row))
        if not data_url:
            return None
        
        # 解析data URL: data:image/png;base64,xxxxx
        match = re.match(r'data:([^;]+);base64,(.+)', data_url)
        if match:
            mime_type = match.group(1)
            base64_data = match.group(2)
            return base64_data, mime_type
        
        return None
    
    def has_images(self) -> bool:
        """检查是否成功提取到图片"""
        return len(self.images) > 0
    
    def get_image_count(self) -> int:
        """获取提取到的图片数量"""
        return len(self.images)


class AIModelProcessor:
    def __init__(self, config_file: str = "config.yaml", providers_file: str = "providers.yaml"):
        """初始化AI模型处理器"""
        self.config = self.load_config(config_file)
        self.providers = self.load_providers(providers_file)
        self.provider_config = self.get_provider_config()
        self.setup_logging()
        self.csv_lock = Lock()  # CSV/Excel文件写入锁
        self.excel_image_extractor: Optional[ExcelImageExtractor] = None  # Excel图片提取器
        
    def load_config(self, config_file: str) -> Dict[str, Any]:
        """加载运行配置文件"""
        default_config = {
            "provider": "openai",
            "model_name": "gpt-4o",
            "temperature": 0.6,
            "max_tokens": 2000,
            "input_file": "sample_data.csv",  # 支持csv和xlsx/xls
            "csv_input_file": "sample_data.csv",  # 向后兼容
            "prompt_file": "system_prompt.md",
            "user_prompt_column": "user_prompt",
            "image_column": "",  # 图片列名（Excel嵌入图片或文件路径）
            "image_source": "auto",  # 图片来源: auto, embedded, path
            "image_base_path": "",
            "image_detail": "auto",
            "max_workers": 3,
            "request_delay": 0.5
        }

        if os.path.exists(config_file):
            with open(config_file, 'r', encoding='utf-8') as f:
                user_config = yaml.safe_load(f)
                default_config.update(user_config)
        else:
            with open(config_file, 'w', encoding='utf-8') as f:
                yaml.dump(default_config, f, allow_unicode=True, default_flow_style=False)
            print(f"📝 已创建默认配置文件: {config_file}")

        # 向后兼容: 如果只设置了csv_input_file而没有input_file，使用csv_input_file
        if "input_file" not in default_config or not default_config.get("input_file"):
            default_config["input_file"] = default_config.get("csv_input_file", "sample_data.csv")
        
        return default_config
    
    def load_providers(self, providers_file: str) -> Dict[str, Any]:
        """加载Provider配置文件"""
        default_providers = {
            "providers": {
                "openai": {
                    "api_url": "https://api.openai.com/v1/chat/completions",
                    "api_key": "",
                    "api_type": "openai",
                    "timeout": 60,
                    "max_retries": 3,
                    "retry_delay": 2
                }
            },
            "default_provider": "openai"
        }

        if os.path.exists(providers_file):
            with open(providers_file, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        else:
            with open(providers_file, 'w', encoding='utf-8') as f:
                yaml.dump(default_providers, f, allow_unicode=True, default_flow_style=False)
            print(f"📝 已创建默认Provider配置文件: {providers_file}")
            return default_providers
    
    def get_provider_config(self) -> Dict[str, Any]:
        """获取当前Provider的配置，支持从环境变量读取API密钥"""
        provider_name = self.config.get("provider", self.providers.get("default_provider", "openai"))
        providers = self.providers.get("providers", {})
        
        if provider_name not in providers:
            print(f"❌ Provider '{provider_name}' 不存在于 providers.yaml")
            print(f"可用的Provider: {', '.join(providers.keys())}")
            sys.exit(1)
        
        config = providers[provider_name].copy()
        
        # 如果配置文件中没有 API 密钥，尝试从环境变量读取
        # 环境变量命名规则: {PROVIDER}_API_KEY (大写)
        # 例如: OPENAI_API_KEY, ANTHROPIC_API_KEY, DEEPSEEK_API_KEY
        if not config.get("api_key"):
            env_key_name = f"{provider_name.upper()}_API_KEY"
            env_api_key = os.environ.get(env_key_name)
            if env_api_key:
                config["api_key"] = env_api_key
                print(f"🔑 已从环境变量 {env_key_name} 读取API密钥")
        
        return config
    
    def setup_logging(self):
        """设置日志"""
        file_handler = logging.FileHandler('ai_processor.log', encoding='utf-8')
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(logging.Formatter('%(message)s'))
        
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        # 避免重复添加handler
        if not self.logger.handlers:
            self.logger.addHandler(file_handler)
            self.logger.addHandler(console_handler)
        
        logging.getLogger('urllib3').setLevel(logging.WARNING)
        logging.getLogger('requests').setLevel(logging.WARNING)
    
    def load_system_prompt(self) -> str:
        """加载系统提示词"""
        prompt_file = self.config["prompt_file"]
        if not os.path.exists(prompt_file):
            self.logger.error(f"❌ 提示词文件不存在: {prompt_file}")
            return ""
        
        with open(prompt_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        if 'system_prompt = """' in content:
            start = content.find('system_prompt = """') + len('system_prompt = """')
            end = content.rfind('"""')
            if end > start:
                content = content[start:end]
        
        return content.strip()
    
    def check_row_processed(self, df: pd.DataFrame, index: int, response_col: str) -> bool:
        """检查指定行是否已经处理过"""
        if response_col not in df.columns:
            return False
        
        response = df.at[index, response_col]
        return not pd.isna(response) and str(response).strip() != ""
    
    def encode_image_to_base64(self, image_path: str) -> Optional[str]:
        """将本地图片转换为Base64编码的data URL"""
        if not os.path.exists(image_path):
            self.logger.error(f"❌ 图片不存在: {image_path}")
            return None
        
        mime_type, _ = mimetypes.guess_type(image_path)
        supported_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        
        if mime_type not in supported_types:
            self.logger.error(f"❌ 不支持的图片格式: {mime_type}")
            return None
        
        try:
            with open(image_path, 'rb') as f:
                image_data = base64.b64encode(f.read()).decode('utf-8')
            return f"data:{mime_type};base64,{image_data}"
        except Exception as e:
            self.logger.error(f"❌ 读取图片失败: {str(e)}")
            return None
    
    def get_image_base64_raw(self, image_path: str) -> Optional[Tuple[str, str]]:
        """获取图片的原始Base64数据和MIME类型"""
        if not os.path.exists(image_path):
            return None
        
        mime_type, _ = mimetypes.guess_type(image_path)
        if not mime_type:
            return None
        
        try:
            with open(image_path, 'rb') as f:
                image_data = base64.b64encode(f.read()).decode('utf-8')
            return image_data, mime_type
        except:
            return None
    
    def is_excel_file(self, file_path: str) -> bool:
        """判断是否为Excel文件"""
        return file_path.lower().endswith(('.xlsx', '.xls', '.xlsm'))
    
    def load_input_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """
        加载输入文件（支持CSV和Excel）
        
        Args:
            file_path: 文件路径
            
        Returns:
            DataFrame或None
        """
        if not os.path.exists(file_path):
            self.logger.error(f"❌ 文件不存在: {file_path}")
            return None
        
        try:
            if self.is_excel_file(file_path):
                if not HAS_OPENPYXL:
                    self.logger.error("❌ 需要安装openpyxl来处理Excel文件: pip install openpyxl")
                    return None
                
                self.logger.info(f"📊 正在加载Excel文件: {file_path}")
                df = pd.read_excel(file_path, engine='openpyxl')
                
                # 初始化Excel图片提取器
                self.excel_image_extractor = ExcelImageExtractor(file_path)
                if self.excel_image_extractor.has_images():
                    self.logger.info(f"🖼️ 从Excel中提取到 {self.excel_image_extractor.get_image_count()} 张嵌入图片")
                
                return df
            else:
                self.logger.info(f"📊 正在加载CSV文件: {file_path}")
                return pd.read_csv(file_path)
                
        except Exception as e:
            self.logger.error(f"❌ 加载文件失败: {str(e)}")
            return None
    
    def get_output_file_path(self, input_file: str) -> str:
        """
        获取输出文件路径
        
        对于Excel文件，输出为新的Excel文件（保留图片）
        对于CSV文件，直接使用原文件
        
        Args:
            input_file: 输入文件路径
            
        Returns:
            输出文件路径
        """
        if self.is_excel_file(input_file):
            # Excel 输入 -> 新 Excel 输出（保留图片）
            base, ext = os.path.splitext(input_file)
            return base + "_results" + ext
        else:
            return input_file
    
    def _copy_excel_with_images(self, input_file: str, output_file: str) -> bool:
        """
        复制Excel文件（保留所有图片和格式）
        
        Args:
            input_file: 源文件路径
            output_file: 目标文件路径
            
        Returns:
            是否成功复制
        """
        import shutil
        try:
            shutil.copy2(input_file, output_file)
            self.logger.info(f"📋 已复制Excel文件到: {output_file}")
            return True
        except Exception as e:
            self.logger.error(f"❌ 复制Excel文件失败: {str(e)}")
            return False
    
    def save_output_file(self, df: pd.DataFrame, input_file: str):
        """
        保存输出文件
        
        对于Excel输入，复制原文件到新文件（保留图片），然后更新数据列
        对于CSV输入，直接保存回原文件
        
        Args:
            df: DataFrame
            input_file: 输入文件路径
        """
        try:
            output_file = self.get_output_file_path(input_file)
            
            if self.is_excel_file(input_file):
                # Excel文件：复制原文件（首次），然后用openpyxl更新数据
                if not os.path.exists(output_file):
                    if not self._copy_excel_with_images(input_file, output_file):
                        return
                
                # 使用openpyxl更新结果列（保留图片）
                self._update_excel_results(df, output_file)
                
                # 首次保存时提示
                if not hasattr(self, '_output_file_logged') or not self._output_file_logged:
                    self.logger.info(f"💾 结果将保存到: {output_file}（保留所有图片）")
                    self._output_file_logged = True
            else:
                # CSV文件：直接保存
                df.to_csv(output_file, index=False)
                
        except Exception as e:
            self.logger.error(f"❌ 保存文件失败: {str(e)}")
    
    def _update_excel_results(self, df: pd.DataFrame, output_file: str):
        """
        使用openpyxl更新Excel文件的结果列（保留图片）
        
        Args:
            df: 包含结果的DataFrame
            output_file: Excel文件路径
        """
        try:
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 获取结果列名
            model_name_safe = self.config["model_name"].replace("-", "_").replace(".", "_")
            response_col = f"ai_response_{model_name_safe}"
            
            if response_col not in df.columns:
                wb.close()
                return
            
            # 查找或创建结果列
            header_row = 1
            result_col_idx = None
            
            # 查找现有列
            for col_idx in range(1, ws.max_column + 2):
                cell_value = ws.cell(row=header_row, column=col_idx).value
                if cell_value == response_col:
                    result_col_idx = col_idx
                    break
                if cell_value is None:
                    # 新列
                    result_col_idx = col_idx
                    ws.cell(row=header_row, column=col_idx, value=response_col)
                    break
            
            if result_col_idx is None:
                result_col_idx = ws.max_column + 1
                ws.cell(row=header_row, column=result_col_idx, value=response_col)
            
            # 写入结果数据
            for idx, value in enumerate(df[response_col]):
                row_num = idx + 2  # Excel行号（跳过标题行）
                if pd.notna(value) and str(value).strip():
                    ws.cell(row=row_num, column=result_col_idx, value=str(value))
            
            wb.save(output_file)
            wb.close()
            
        except Exception as e:
            self.logger.error(f"❌ 更新Excel结果列失败: {str(e)}")
    
    def get_image_for_row(self, row_index: int, row: pd.Series, image_col: str) -> Optional[str]:
        """
        获取指定行的图片（支持嵌入图片和文件路径）
        
        Args:
            row_index: DataFrame行索引
            row: 行数据
            image_col: 图片列名
            
        Returns:
            图片的Base64 data URL，或图片文件路径
        """
        image_source = self.config.get("image_source", "auto")
        
        # 如果有Excel图片提取器且是嵌入图片模式
        if self.excel_image_extractor and image_source in ("auto", "embedded"):
            # Excel的行号 = DataFrame索引 + 2 (索引从0开始，Excel行号从1开始且有标题行)
            excel_row = row_index + 2
            base64_image = self.excel_image_extractor.get_image_base64(excel_row)
            if base64_image:
                return base64_image
        
        # 检查是否有文件路径
        if image_col and image_col in row.index:
            img_value = row.get(image_col, "")
            if pd.notna(img_value) and str(img_value).strip():
                img_path = str(img_value).strip()
                # 如果是完整的data URL，直接返回
                if img_path.startswith("data:"):
                    return img_path
                # 否则作为文件路径处理
                return img_path
        
        return None
    
    def build_user_message_openai(self, text: str, image_data: str = None) -> Union[str, List]:
        """
        构建OpenAI格式的用户消息
        
        Args:
            text: 用户文本
            image_data: 图片数据，可以是:
                - Base64 data URL (data:image/xxx;base64,...)
                - 本地文件路径
        """
        if not image_data:
            return text
        
        # 判断是data URL还是文件路径
        if image_data.startswith("data:"):
            # 已经是Base64 data URL
            image_url = image_data
        else:
            # 是文件路径，需要转换
            image_base_path = self.config.get("image_base_path", "")
            if image_base_path and not os.path.isabs(image_data):
                image_data = os.path.join(image_base_path, image_data)
            
            image_url = self.encode_image_to_base64(image_data)
            if not image_url:
                return text
        
        content = []
        if text and text.strip():
            content.append({"type": "text", "text": text})
        
        content.append({
            "type": "image_url",
            "image_url": {
                "url": image_url,
                "detail": self.config.get("image_detail", "auto")
            }
        })
        
        return content
    
    def build_user_message_anthropic(self, text: str, image_data: str = None) -> List:
        """
        构建Anthropic格式的用户消息
        
        Args:
            text: 用户文本
            image_data: 图片数据，可以是:
                - Base64 data URL (data:image/xxx;base64,...)
                - 本地文件路径
        """
        content = []
        
        if image_data:
            # 判断是data URL还是文件路径
            if image_data.startswith("data:"):
                # 解析data URL
                match = re.match(r'data:([^;]+);base64,(.+)', image_data)
                if match:
                    mime_type = match.group(1)
                    base64_data = match.group(2)
                    content.append({
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": mime_type,
                            "data": base64_data
                        }
                    })
            else:
                # 是文件路径
                image_base_path = self.config.get("image_base_path", "")
                if image_base_path and not os.path.isabs(image_data):
                    image_data = os.path.join(image_base_path, image_data)
                
                result = self.get_image_base64_raw(image_data)
                if result:
                    base64_data, mime_type = result
                    content.append({
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": mime_type,
                            "data": base64_data
                        }
                    })
        
        if text and text.strip():
            content.append({"type": "text", "text": text})
        
        return content if content else [{"type": "text", "text": text or ""}]
    
    def call_api_openai(self, user_prompt: str, system_prompt: str, image_path: str = None) -> Optional[str]:
        """调用OpenAI兼容格式的API"""
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.provider_config['api_key']}"
        }
        
        user_content = self.build_user_message_openai(user_prompt, image_path)
        
        data = {
            "model": self.config["model_name"],
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            "temperature": self.config.get("temperature", 0.6)
        }
        
        if "max_tokens" in self.config:
            data["max_tokens"] = self.config["max_tokens"]
        
        max_retries = self.provider_config.get("max_retries", 3)
        retry_delay = self.provider_config.get("retry_delay", 1)
        
        for attempt in range(max_retries):
            try:
                response = requests.post(
                    self.provider_config["api_url"],
                    headers=headers,
                    json=data,
                    timeout=self.provider_config.get("timeout", 60)
                )
                
                if response.status_code == 200:
                    result = response.json()
                    if "choices" in result and len(result["choices"]) > 0:
                        return result["choices"][0]["message"]["content"]
                else:
                    self.logger.error(f"❌ API调用失败 (状态码: {response.status_code})")
                
            except requests.exceptions.RequestException as e:
                if attempt == max_retries - 1:
                    self.logger.error(f"❌ API调用失败: {str(e)[:50]}...")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (attempt + 1))
        
        return None
    
    def call_api_anthropic(self, user_prompt: str, system_prompt: str, image_path: str = None) -> Optional[str]:
        """调用Anthropic Claude API"""
        headers = {
            "Content-Type": "application/json",
            "x-api-key": self.provider_config['api_key'],
            "anthropic-version": self.provider_config.get("api_version", "2023-06-01")
        }
        
        user_content = self.build_user_message_anthropic(user_prompt, image_path)
        
        data = {
            "model": self.config["model_name"],
            "max_tokens": self.config.get("max_tokens", 4096),
            "system": system_prompt,
            "messages": [
                {"role": "user", "content": user_content}
            ]
        }
        
        if "temperature" in self.config:
            data["temperature"] = self.config["temperature"]
        
        max_retries = self.provider_config.get("max_retries", 3)
        retry_delay = self.provider_config.get("retry_delay", 2)
        
        for attempt in range(max_retries):
            try:
                response = requests.post(
                    self.provider_config["api_url"],
                    headers=headers,
                    json=data,
                    timeout=self.provider_config.get("timeout", 60)
                )
                
                if response.status_code == 200:
                    result = response.json()
                    if "content" in result and len(result["content"]) > 0:
                        return result["content"][0]["text"]
                else:
                    self.logger.error(f"❌ API调用失败 (状态码: {response.status_code})")
                
            except requests.exceptions.RequestException as e:
                if attempt == max_retries - 1:
                    self.logger.error(f"❌ API调用失败: {str(e)[:50]}...")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (attempt + 1))
        
        return None
    
    def call_api_google(self, user_prompt: str, system_prompt: str, image_data: str = None) -> Optional[str]:
        """
        调用Google Gemini API
        
        Args:
            user_prompt: 用户提示词
            system_prompt: 系统提示词
            image_data: 图片数据，可以是Base64 data URL或文件路径
        """
        model_name = self.config["model_name"]
        api_key = self.provider_config['api_key']
        base_url = self.provider_config["api_url"]
        url = f"{base_url}/models/{model_name}:generateContent?key={api_key}"
        
        headers = {"Content-Type": "application/json"}
        
        # 构建内容
        parts = []
        
        # 添加系统提示词作为文本的一部分
        if system_prompt:
            parts.append({"text": f"System: {system_prompt}\n\nUser: {user_prompt}"})
        else:
            parts.append({"text": user_prompt})
        
        # 添加图片
        if image_data:
            # 判断是data URL还是文件路径
            if image_data.startswith("data:"):
                # 解析data URL
                match = re.match(r'data:([^;]+);base64,(.+)', image_data)
                if match:
                    mime_type = match.group(1)
                    base64_data = match.group(2)
                    parts.append({
                        "inline_data": {
                            "mime_type": mime_type,
                            "data": base64_data
                        }
                    })
            else:
                # 是文件路径
                image_base_path = self.config.get("image_base_path", "")
                if image_base_path and not os.path.isabs(image_data):
                    image_data = os.path.join(image_base_path, image_data)
                
                result = self.get_image_base64_raw(image_data)
                if result:
                    base64_data, mime_type = result
                    parts.append({
                        "inline_data": {
                            "mime_type": mime_type,
                            "data": base64_data
                        }
                    })
        
        data = {
            "contents": [{"parts": parts}],
            "generationConfig": {
                "temperature": self.config.get("temperature", 0.6),
                "maxOutputTokens": self.config.get("max_tokens", 2048)
            }
        }
        
        max_retries = self.provider_config.get("max_retries", 3)
        retry_delay = self.provider_config.get("retry_delay", 2)
        
        for attempt in range(max_retries):
            try:
                response = requests.post(
                    url,
                    headers=headers,
                    json=data,
                    timeout=self.provider_config.get("timeout", 60)
                )
                
                if response.status_code == 200:
                    result = response.json()
                    if "candidates" in result and len(result["candidates"]) > 0:
                        candidate = result["candidates"][0]
                        if "content" in candidate and "parts" in candidate["content"]:
                            return candidate["content"]["parts"][0]["text"]
                else:
                    self.logger.error(f"❌ API调用失败 (状态码: {response.status_code})")
                
            except requests.exceptions.RequestException as e:
                if attempt == max_retries - 1:
                    self.logger.error(f"❌ API调用失败: {str(e)[:50]}...")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (attempt + 1))
        
        return None
    
    def call_ai_api(self, user_prompt: str, system_prompt: str, image_path: str = None) -> Optional[Dict[str, Any]]:
        """统一的API调用入口，根据api_type选择调用方式"""
        api_type = self.provider_config.get("api_type", "openai")
        
        if api_type == "openai":
            content = self.call_api_openai(user_prompt, system_prompt, image_path)
        elif api_type == "anthropic":
            content = self.call_api_anthropic(user_prompt, system_prompt, image_path)
        elif api_type == "google":
            content = self.call_api_google(user_prompt, system_prompt, image_path)
        else:
            self.logger.error(f"❌ 不支持的API类型: {api_type}")
            return None
        
        if content:
            return self.parse_ai_response(content)
        return None
    
    def parse_ai_response(self, content: str) -> Optional[Dict[str, Any]]:
        """解析AI返回的JSON内容"""
        try:
            if content.strip().startswith('{') and content.strip().endswith('}'):
                return json.loads(content)
            
            if '```json' in content:
                start = content.find('```json') + 7
                end = content.find('```', start)
                if end > start:
                    json_content = content[start:end].strip()
                    return json.loads(json_content)
            
            start = content.find('{')
            end = content.rfind('}') + 1
            if start >= 0 and end > start:
                json_content = content[start:end]
                return json.loads(json_content)
            
            self.logger.error(f"❌ 无法解析AI响应为JSON")
            return None
            
        except json.JSONDecodeError as e:
            self.logger.error(f"❌ JSON解析错误: {str(e)[:30]}...")
            return None
    
    def process_single_row(self, index: int, user_prompt: str, system_prompt: str,
                          df: pd.DataFrame, response_col: str,
                          image_path: str = None) -> bool:
        """处理单行数据（线程安全）"""
        try:
            time.sleep(self.config.get("request_delay", 0.5))

            result = self.call_ai_api(user_prompt, system_prompt, image_path)

            if result:
                # 直接保存完整的 AI 响应（JSON 格式）
                full_response = json.dumps(result, ensure_ascii=False)
                with self.csv_lock:
                    df.at[index, response_col] = full_response
                return True
            else:
                return False

        except Exception as e:
            return False
    
    def process_csv(self) -> bool:
        """处理CSV文件（向后兼容）"""
        return self.process_file()
    
    def process_file(self) -> bool:
        """处理输入文件（支持CSV和Excel）"""
        # 优先使用input_file，向后兼容csv_input_file
        input_file = self.config.get("input_file") or self.config.get("csv_input_file")
        
        if not input_file:
            self.logger.error("❌ 未配置输入文件")
            return False
        
        # 加载输入文件（用于获取原始数据和图片）
        df = self.load_input_file(input_file)
        if df is None:
            return False
        
        user_prompt_col = self.config["user_prompt_column"]
        
        # 检查用户提示列是否存在，如果不存在则创建空列
        if user_prompt_col not in df.columns:
            self.logger.warning(f"⚠️ 文件中不存在列 '{user_prompt_col}'，将创建空列")
            df[user_prompt_col] = ""
        
        # 显示当前使用的Provider和模型
        provider_name = self.config.get("provider", "unknown")
        model_name = self.config.get("model_name", "unknown")
        api_type = self.provider_config.get("api_type", "unknown")
        self.logger.info(f"🤖 Provider: {provider_name} | 模型: {model_name} | API类型: {api_type}")
        
        image_col = self.config.get("image_column", "")
        has_image_col = image_col and image_col in df.columns
        
        # 检查是否有Excel嵌入图片
        has_embedded_images = self.excel_image_extractor and self.excel_image_extractor.has_images()
        
        if image_col and image_col not in df.columns and not has_embedded_images:
            self.logger.warning(f"⚠️ 配置的图片列 '{image_col}' 不存在，将使用纯文本模式")
            has_image_col = False
        
        if has_embedded_images:
            self.logger.info(f"🖼️ 已启用嵌入图片模式，共 {self.excel_image_extractor.get_image_count()} 张图片")
        elif has_image_col:
            self.logger.info(f"🖼️ 已启用图片路径模式，图片列: {image_col}")
        
        system_prompt = self.load_system_prompt()
        if not system_prompt:
            self.logger.error("❌ 无法加载系统提示词")
            return False
        
        # 使用单一的响应列，不再拆分 reasoning 和 classification
        model_name_safe = self.config["model_name"].replace("-", "_").replace(".", "_")
        response_col = f"ai_response_{model_name_safe}"
        
        # 对于Excel文件，尝试加载已有的结果文件进行断点续传
        output_file = self.get_output_file_path(input_file)
        if self.is_excel_file(input_file) and os.path.exists(output_file):
            try:
                # 加载已有的结果Excel文件
                existing_df = pd.read_excel(output_file, engine='openpyxl')
                # 将已处理的结果合并到当前df
                if response_col in existing_df.columns:
                    df[response_col] = existing_df[response_col]
                    self.logger.info(f"📂 已加载之前的结果文件: {output_file}")
            except Exception as e:
                self.logger.warning(f"⚠️ 无法加载结果文件，将重新开始: {str(e)}")
        
        if response_col not in df.columns:
            df[response_col] = ""
        
        total_rows = len(df)
        rows_to_process = []
        processed_count = 0
        
        file_type = "Excel" if self.is_excel_file(input_file) else "CSV"
        self.logger.info(f"📊 扫描{file_type}文件，检查处理状态...")
        
        for index, row in df.iterrows():
            if self.check_row_processed(df, index, response_col):
                processed_count += 1
                continue
            
            # 构建用户提示词
            col_value = str(row[user_prompt_col]) if pd.notna(row[user_prompt_col]) else ""
            user_prompt_template = self.config.get("user_prompt_template", "")
            if user_prompt_template and col_value:
                user_prompt = user_prompt_template.format(col_value)
            else:
                user_prompt = col_value
            
            # 获取图片数据（支持嵌入图片和文件路径）
            image_data = self.get_image_for_row(index, row, image_col)
            
            rows_to_process.append((index, user_prompt, image_data))
        
        self.logger.info(f"📈 扫描完成: 总计 {total_rows} 行，已处理 {processed_count} 行，待处理 {len(rows_to_process)} 行")
        
        if not rows_to_process:
            self.logger.info("✅ 所有数据已处理完成")
            return True
        
        self.logger.info(f"🚀 开始处理 {len(rows_to_process)} 条数据 (线程数: {self.config['max_workers']})")
        
        new_processed_count = 0
        max_workers = self.config.get("max_workers", 3)
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            with tqdm(total=len(rows_to_process), desc="📊 处理进度", 
                     bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]",
                     ncols=80) as pbar:
                future_to_index = {}
                for index, user_prompt, image_data in rows_to_process:
                    future = executor.submit(
                        self.process_single_row, 
                        index, user_prompt, system_prompt, 
                        df, response_col,
                        image_data
                    )
                    future_to_index[future] = index
                
                for future in as_completed(future_to_index):
                    index = future_to_index[future]
                    try:
                        success = future.result()
                        if success:
                            new_processed_count += 1
                        
                        if new_processed_count % 10 == 0:
                            with self.csv_lock:
                                self.save_output_file(df, input_file)
                                
                        pbar.update(1)
                        
                    except Exception as e:
                        pbar.update(1)
        
        with self.csv_lock:
            self.save_output_file(df, input_file)
        
        self.logger.info(f"🎉 处理完成！共处理 {new_processed_count} 条新数据")
        return True
    
    def reset_progress(self):
        """重置进度"""
        input_file = self.config.get("input_file") or self.config.get("csv_input_file")
        if not os.path.exists(input_file):
            self.logger.error(f"❌ 文件不存在: {input_file}")
            return
        
        # 对于Excel文件，删除结果文件
        if self.is_excel_file(input_file):
            output_file = self.get_output_file_path(input_file)
            if os.path.exists(output_file):
                os.remove(output_file)
                self.logger.info(f"🔄 已删除结果文件: {output_file}")
            else:
                self.logger.info("🔄 没有找到需要重置的结果文件")
        else:
            # CSV文件直接清空结果列
            df = self.load_input_file(input_file)
            if df is None:
                return
            
            model_name_safe = self.config["model_name"].replace("-", "_").replace(".", "_")
            response_col = f"ai_response_{model_name_safe}"
            
            if response_col in df.columns:
                df[response_col] = ""
            
            self.save_output_file(df, input_file)
            self.logger.info("🔄 进度已重置，已清空所有处理结果")
    
    def show_status(self):
        """显示当前状态"""
        input_file = self.config.get("input_file") or self.config.get("csv_input_file")
        if not os.path.exists(input_file):
            print(f"❌ 文件不存在: {input_file}")
            return
        
        # 加载输入文件获取总行数和图片信息
        df = self.load_input_file(input_file)
        if df is None:
            return
        
        total_rows = len(df)
        model_name_safe = self.config["model_name"].replace("-", "_").replace(".", "_")
        response_col = f"ai_response_{model_name_safe}"
        
        # 对于Excel文件，检查是否存在结果CSV文件
        output_file = self.get_output_file_path(input_file)
        processed_rows = 0
        
        if self.is_excel_file(input_file) and os.path.exists(output_file):
            # 从结果Excel文件读取处理状态
            try:
                result_df = pd.read_excel(output_file, engine='openpyxl')
                for index in range(len(result_df)):
                    if self.check_row_processed(result_df, index, response_col):
                        processed_rows += 1
            except Exception:
                pass
        else:
            # CSV文件直接从原文件检查
            for index in range(total_rows):
                if self.check_row_processed(df, index, response_col):
                    processed_rows += 1
        
        progress_pct = processed_rows/total_rows*100 if total_rows > 0 else 0
        remaining = total_rows - processed_rows
        
        provider_name = self.config.get("provider", "unknown")
        model_name = self.config.get("model_name", "unknown")
        file_type = "Excel" if self.is_excel_file(input_file) else "CSV"
        
        print(f"\n📊 处理状态")
        print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"📁 文件类型:   {file_type}")
        print(f"📄 输入文件:   {input_file}")
        if output_file != input_file:
            print(f"💾 输出文件:   {output_file}")
        print(f"🤖 Provider:   {provider_name}")
        print(f"📦 模型:       {model_name}")
        print(f"📝 总行数:     {total_rows:,}")
        print(f"✅ 已处理:     {processed_rows:,}")
        print(f"⏳ 待处理:     {remaining:,}")
        print(f"📈 完成率:     {progress_pct:.1f}%")
        print(f"🔧 线程数:     {self.config.get('max_workers', 3)}")
        
        # 显示嵌入图片信息
        if self.excel_image_extractor and self.excel_image_extractor.has_images():
            print(f"🖼️ 嵌入图片:   {self.excel_image_extractor.get_image_count()} 张")
        
        bar_length = 30
        filled_length = int(bar_length * progress_pct / 100)
        bar = "█" * filled_length + "░" * (bar_length - filled_length)
        print(f"📊 进度条:     [{bar}] {progress_pct:.1f}%")
        
        if processed_rows > 0 and remaining > 0:
            avg_time_per_item = 3.0 / self.config.get('max_workers', 3)
            estimated_hours = (remaining * avg_time_per_item) / 3600
            if estimated_hours < 1:
                estimated_minutes = (remaining * avg_time_per_item) / 60
                print(f"⏰ 预估时间:   {estimated_minutes:.0f} 分钟")
            else:
                print(f"⏰ 预估时间:   {estimated_hours:.1f} 小时")
        print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    
    def list_providers(self):
        """列出所有可用的Provider"""
        providers = self.providers.get("providers", {})
        default_provider = self.providers.get("default_provider", "")
        
        print(f"\n📋 可用的Provider列表")
        print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        
        for name, config in providers.items():
            is_default = " (默认)" if name == default_provider else ""
            api_type = config.get("api_type", "unknown")
            # 同时检查配置文件和环境变量
            env_key_name = f"{name.upper()}_API_KEY"
            has_key_in_config = bool(config.get("api_key"))
            has_key_in_env = bool(os.environ.get(env_key_name))
            has_key = "✅" if (has_key_in_config or has_key_in_env) else "❌"
            if has_key_in_env and not has_key_in_config:
                has_key = "✅ (环境变量)"
            models = config.get("available_models", [])
            
            print(f"\n🔹 {name}{is_default}")
            print(f"   API类型: {api_type}")
            print(f"   密钥状态: {has_key}")
            print(f"   可用模型: {', '.join(models[:3])}{'...' if len(models) > 3 else ''}")
        
        print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")


def main():
    parser = argparse.ArgumentParser(description='AI模型调用脚本（支持多Provider）')
    parser.add_argument('--config', default='config.yaml', help='配置文件路径')
    parser.add_argument('--providers', default='providers.yaml', help='Provider配置文件路径')
    parser.add_argument('--reset', action='store_true', help='重置进度')
    parser.add_argument('--status', action='store_true', help='显示状态')
    parser.add_argument('--list-providers', action='store_true', help='列出所有Provider')
    parser.add_argument('--workers', type=int, help='并发线程数量')
    parser.add_argument('--provider', type=str, help='指定使用的Provider')
    parser.add_argument('--model', type=str, help='指定使用的模型')
    
    args = parser.parse_args()
    
    processor = AIModelProcessor(args.config, args.providers)
    
    # 命令行参数覆盖
    if args.workers is not None:
        processor.config["max_workers"] = args.workers
        print(f"🔧 使用命令行指定的线程数: {args.workers}")
    
    if args.provider is not None:
        processor.config["provider"] = args.provider
        processor.provider_config = processor.get_provider_config()
        print(f"🔧 使用命令行指定的Provider: {args.provider}")
    
    if args.model is not None:
        processor.config["model_name"] = args.model
        print(f"🔧 使用命令行指定的模型: {args.model}")
    
    if args.list_providers:
        processor.list_providers()
        return
    
    if args.reset:
        processor.reset_progress()
        return
    
    if args.status:
        processor.show_status()
        return
    
    # 检查API密钥
    if not processor.provider_config.get("api_key"):
        provider_name = processor.config.get("provider", "unknown")
        print(f"⚠️  请在 providers.yaml 中为 '{provider_name}' 设置API密钥")
        return
    
    processor.process_csv()


if __name__ == "__main__":
    main()
